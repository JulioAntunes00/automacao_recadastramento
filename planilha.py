import os
from openpyxl import load_workbook

def atualizar_status_excel(caminho_ano, nome_pasta_pdf, nome_pessoa, observacao):

    
    pasta_excel_raiz = os.path.join(caminho_ano, "PLANILHA DE CONTROLE")
    
    arquivo_excel_encontrado = None
    
    # Lista todos os arquivos na pasta de planilhas
    if not os.path.exists(pasta_excel_raiz):
        print("✖ Erro: Pasta 'PLANILHA DE CONTROLE' não encontrada.")
        return False

    arquivos_existentes = os.listdir(pasta_excel_raiz)

    meses_possiveis = ["JANEIRO", "FEVEREIRO", "MARÇO", "ABRIL", "MAIO", "JUNHO", 
                       "JULHO", "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO", "DEZEMBRO"]
    
    mes_identificado = ""
    for mes in meses_possiveis:
        if mes in nome_pasta_pdf.upper():
            mes_identificado = mes
            break
            
    if not mes_identificado:
        print(f"✖ Erro: Não consegui identificar o mês na pasta '{nome_pasta_pdf}'.")
        return False

    for arquivo in arquivos_existentes:
        if mes_identificado in arquivo.upper() and arquivo.endswith(".xlsx"):
            arquivo_excel_encontrado = arquivo
            break
    
    if not arquivo_excel_encontrado:
        print(f"✖ Erro: Nenhuma planilha encontrada contendo '{mes_identificado}'.")
        return False

    caminho_completo_excel = os.path.join(pasta_excel_raiz, arquivo_excel_encontrado)
    print(f"📂 Abrindo planilha: {arquivo_excel_encontrado}...")

    #Editar o arquivo
    try:
        wb = load_workbook(caminho_completo_excel)
        
        pessoa_encontrada = False

        for nome_aba in wb.sheetnames:
            ws = wb[nome_aba]
            
            for linha in ws.iter_rows(min_row=2): # Começa na linha 2 (pula cabeçalho)
                celula_nome = linha[1]
                
                # Verificamos se a célula não está vazia e se é a pessoa
                if celula_nome.value and str(celula_nome.value).strip().upper() == nome_pessoa:
                                        
                    texto_obs = observacao.strip() if observacao else "VIA WHATS SEM ALTERAÇÃO"
                    
                    linha[6].value = texto_obs  
                    linha[8].value = "OK"       
                    
                    pessoa_encontrada = True
                    print(f"✓ Atualizado na aba '{nome_aba}': {texto_obs}")
                    break
            
            if pessoa_encontrada:
                break

        if pessoa_encontrada:
            wb.save(caminho_completo_excel)
            print("💾 Planilha salva com sucesso!")
            return True
        else:
            print(f"⚠ Aviso: '{nome_pessoa}' não foi encontrado(a) na planilha.")
            return False

    except PermissionError:
        print("✖ ERRO CRÍTICO: A planilha está aberta no Excel!")
        print("➜ Por favor, FECHE o arquivo Excel e tente novamente.")
        return False
    except Exception as e:
        print(f"✖ Erro inesperado ao mexer no Excel: {e}")
        return False